from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


HEADERS = [
    "timestamp",
    "invoice_no",
    "customer_name",
    "subtotal_ex_vat",
    "vat_amount",
    "total_inc_vat",
    "currency",
]


def _ensure_headers(ws) -> None:
    if ws.max_row < 1:
        ws.append(HEADERS)
        return
    row = [cell.value for cell in ws[1]]
    if any(row) and row[: len(HEADERS)] == HEADERS:
        return
    if not any(row):
        ws.append(HEADERS)


def _normalize_invoice_no(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text.isdigit():
        return str(int(text))
    return text


def _format_total_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        return f"{Decimal(text):.2f}"
    except (InvalidOperation, ValueError):
        return text


def write_ledger_xlsx(path: str, row: dict) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)

    if target.exists():
        workbook = load_workbook(target)
        sheet = workbook.active
        _ensure_headers(sheet)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(HEADERS)

    sheet.append([row.get(header) for header in HEADERS])
    workbook.save(target)


def read_invoice_totals(path: str | Path, invoice_no: str) -> dict[str, str] | None:
    target = Path(path)
    if not target.exists():
        return None

    workbook = load_workbook(target, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return None

        headers = [str(value).strip() if value is not None else "" for value in header_row]
        try:
            invoice_idx = headers.index("invoice_no")
            subtotal_idx = headers.index("subtotal_ex_vat")
            vat_idx = headers.index("vat_amount")
            total_idx = headers.index("total_inc_vat")
            currency_idx = headers.index("currency")
        except ValueError:
            return None

        normalized_invoice_no = _normalize_invoice_no(invoice_no)
        latest_row: tuple[Any, ...] | None = None
        for row in rows:
            if not row:
                continue
            row_invoice_no = _normalize_invoice_no(
                row[invoice_idx] if invoice_idx < len(row) else None
            )
            if row_invoice_no == normalized_invoice_no:
                latest_row = row

        if latest_row is None:
            return None

        currency_value = latest_row[currency_idx] if currency_idx < len(latest_row) else None
        currency = str(currency_value).strip() if currency_value is not None else ""
        return {
            "invoice_no": _normalize_invoice_no(
                latest_row[invoice_idx] if invoice_idx < len(latest_row) else invoice_no
            ),
            "subtotal": _format_total_value(
                latest_row[subtotal_idx] if subtotal_idx < len(latest_row) else None
            ),
            "vat": _format_total_value(
                latest_row[vat_idx] if vat_idx < len(latest_row) else None
            ),
            "total": _format_total_value(
                latest_row[total_idx] if total_idx < len(latest_row) else None
            ),
            "currency": currency or "EUR",
        }
    finally:
        workbook.close()
